import pandas as pd
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from transformers import BertTokenizer, BertForSequenceClassification, AdamW
from torch.utils.data import Dataset, DataLoader
import torch
import numpy as np

# 1. Load Excel file using openpyxl
excel_file = '/mnt/data/POFMA Master Dataset - Copy.xlsx'
workbook = load_workbook(excel_file)
sheet = workbook.active

# 2. Extract the relevant columns (Falsehood Claims: T and U; Actor labels: N to S)
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # Skipping the header row
    falsehood_claim = f"{row[19]} {row[20]}"  # Columns T (index 19) and U (index 20) as input
    actors = list(row[13:19])  # Columns N (index 13) to S (index 18) as labels
    actors = [1 if actor == '1' else 0 for actor in actors]  # Convert actor values to binary
    data.append([falsehood_claim, actors])

# Convert to DataFrame
df = pd.DataFrame(data, columns=['Falsehood_Claim', 'Label'])

# 3. Split data into input text and labels
df['Label'] = df['Label'].apply(lambda x: x)  # Labels are already multi-label lists

# 4. Define Dataset class
class POFMADataset(Dataset):
    def __init__(self, texts, labels, tokenizer, max_len=512):
        self.texts = texts
        self.labels = labels
        self.tokenizer = tokenizer
        self.max_len = max_len

    def __len__(self):
        return len(self.texts)

    def __getitem__(self, idx):
        text = self.texts[idx]
        label = self.labels[idx]
        encoding = self.tokenizer(text, return_tensors='pt', padding='max_length', truncation=True, max_length=self.max_len)
        input_ids = encoding['input_ids'].squeeze()
        attention_mask = encoding['attention_mask'].squeeze()
        return {
            'input_ids': input_ids,
            'attention_mask': attention_mask,
            'labels': torch.tensor(label, dtype=torch.float)
        }

# 5. Tokenizer and model
tokenizer = BertTokenizer.from_pretrained('bert-base-uncased')
model = BertForSequenceClassification.from_pretrained('bert-base-uncased', num_labels=6)  # Assuming 6 labels (multi-label classification)

# 6. Split the data into train and validation sets
train_texts, val_texts, train_labels, val_labels = train_test_split(df['Falsehood_Claim'], df['Label'], test_size=0.2, random_state=42)

# 7. Prepare Dataset and DataLoader
train_dataset = POFMADataset(train_texts.values, train_labels.values, tokenizer)
val_dataset = POFMADataset(val_texts.values, val_labels.values, tokenizer)

train_loader = DataLoader(train_dataset, batch_size=8, shuffle=True)
val_loader = DataLoader(val_dataset, batch_size=8, shuffle=False)

# 8. Define Training and Evaluation Functions
def train_epoch(model, dataloader, optimizer, device):
    model = model.train()
    total_loss = 0
    for batch in dataloader:
        optimizer.zero_grad()
        input_ids = batch['input_ids'].to(device)
        attention_mask = batch['attention_mask'].to(device)
        labels = batch['labels'].to(device)

        outputs = model(input_ids=input_ids, attention_mask=attention_mask, labels=labels)
        loss = outputs.loss
        total_loss += loss.item()
        loss.backward()
        optimizer.step()

    return total_loss / len(dataloader)

def eval_model(model, dataloader, device):
    model = model.eval()
    total_loss = 0
    all_predictions = []
    all_labels = []

    with torch.no_grad():
        for batch in dataloader:
            input_ids = batch['input_ids'].to(device)
            attention_mask = batch['attention_mask'].to(device)
            labels = batch['labels'].to(device)

            outputs = model(input_ids=input_ids, attention_mask=attention_mask)
            logits = outputs.logits

            # Sigmoid activation for multi-label classification
            predictions = torch.sigmoid(logits)
            rounded_preds = (predictions > 0.5).int()

            all_predictions.append(rounded_preds.cpu().numpy())
            all_labels.append(labels.cpu().numpy())

            loss = outputs.loss
            total_loss += loss.item()

    all_predictions = np.concatenate(all_predictions, axis=0)
    all_labels = np.concatenate(all_labels, axis=0)

    # Calculate metrics
    accuracy = accuracy_score(all_labels, all_predictions)
    precision = precision_score(all_labels, all_predictions, average='micro')
    recall = recall_score(all_labels, all_predictions, average='micro')
    f1 = f1_score(all_labels, all_predictions, average='micro')

    return total_loss / len(dataloader), accuracy, precision, recall, f1

# 9. Set Device, Optimizer, and Train
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model = model.to(device)
optimizer = AdamW(model.parameters(), lr=5e-5)

# 10. Training Loop with Evaluation
epochs = 3
for epoch in range(epochs):
    # Train the model
    train_loss = train_epoch(model, train_loader, optimizer, device)
    
    # Evaluate the model
    val_loss, accuracy, precision, recall, f1 = eval_model(model, val_loader, device)
    
    # Print the metrics for each epoch
    print(f'Epoch {epoch+1}/{epochs}')
    print(f'Train Loss: {train_loss:.3f}')
    print(f'Validation Loss: {val_loss:.3f}')
    print(f'Accuracy: {accuracy:.4f}')
    print(f'Precision: {precision:.4f}')
    print(f'Recall: {recall:.4f}')
    print(f'F1 Score: {f1:.4f}')

# Save the model after training
model.save_pretrained('pofma_model')
tokenizer.save_pretrained('pofma_tokenizer')
